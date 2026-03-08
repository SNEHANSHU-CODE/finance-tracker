"""
Document Service
Handles decoding base64 vault documents (PDF, CSV, XLSX),
extracting text, and chunking into RAG-ready pieces.

Rename from pdfService.py → documentService.py.
Update import in ragPipeline.py accordingly.
"""
import base64
import csv
import io
import logging
from dataclasses import dataclass
from typing import List, Optional

import fitz  # pymupdf — PDF only
from app.utils.PdfOcrService import PDFOCRService, OCRError
import openpyxl  # XLSX
from langchain_text_splitters import RecursiveCharacterTextSplitter

from app.core.config import settings

logger = logging.getLogger(__name__)

# Supported MIME types — must match vaultModel allowlist
MIME_PDF  = "application/pdf"
MIME_CSV  = "text/csv"
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SUPPORTED_MIME_TYPES = {MIME_PDF, MIME_CSV, MIME_XLSX}

# Minimum characters per page from pymupdf before we consider it a scanned page
# and fall back to OCR for that specific page only.
OCR_FALLBACK_MIN_CHARS_PER_PAGE = 50


@dataclass
class TextChunk:
    """A single chunk ready for embedding."""
    text: str
    chunk_index: int
    page_number: Optional[int]  # page for PDF; sheet index for spreadsheets; None for CSV
    source: str                 # original filename


class DocumentService:
    """
    Stateless multi-format document processor.
    Dispatches to the correct extractor based on mimeType,
    then chunks all formats the same way via langchain splitter.

    Dependencies (all in requirements.txt): pymupdf, openpyxl
    """

    _splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.CHUNK_SIZE,
        chunk_overlap=settings.CHUNK_OVERLAP,
        length_function=len,
        separators=["\n\n", "\n", ". ", " ", ""],
    )

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    @classmethod
    def process_vault_document(
        cls,
        base64_data: str,
        original_name: str,
        mime_type: str = MIME_PDF,
    ) -> List[TextChunk]:
        """
        Main entry point — replaces PDFService.process_vault_document().
        Accepts PDF, CSV, XLSX.

        Args:
            base64_data:   Vault.data field (base64-encoded file bytes)
            original_name: Vault.originalName (used as source label)
            mime_type:     Vault.mimeType — controls which extractor runs

        Returns:
            List of TextChunk objects ready for embedding
        """
        if mime_type not in SUPPORTED_MIME_TYPES:
            raise ValueError(
                f"Unsupported mimeType '{mime_type}'. "
                f"Supported: {SUPPORTED_MIME_TYPES}"
            )

        try:
            raw_bytes = cls._decode_base64(base64_data)

            if mime_type == MIME_PDF:
                pages = cls._extract_pdf(raw_bytes, original_name)
            elif mime_type == MIME_CSV:
                pages = cls._extract_csv(raw_bytes, original_name)
            elif mime_type == MIME_XLSX:
                pages = cls._extract_xlsx(raw_bytes, original_name)

            chunks = cls._chunk_pages(pages, original_name)
            logger.info(
                "Processed '%s' (%s) → %d sections → %d chunks",
                original_name, mime_type, len(pages), len(chunks),
            )
            return chunks

        except Exception as e:
            logger.error("Failed to process '%s' (%s): %s", original_name, mime_type, e)
            raise

    # ------------------------------------------------------------------
    # Base64 decode
    # ------------------------------------------------------------------

    @staticmethod
    def _decode_base64(base64_data: str) -> bytes:
        """Strip data-URI prefix if present, then decode."""
        if "," in base64_data:
            base64_data = base64_data.split(",", 1)[1]
        try:
            return base64.b64decode(base64_data)
        except Exception as e:
            raise ValueError(f"Invalid base64 data: {e}") from e

    # ------------------------------------------------------------------
    # Extractors — each returns List[{page_number, text}]
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_pdf(pdf_bytes: bytes, source: str) -> List[dict]:
        """
        Extract text per page from a PDF.

        Strategy (per page):
          1. Try pymupdf first — fast, free, works for text-based PDFs.
          2. If a page yields fewer than OCR_FALLBACK_MIN_CHARS_PER_PAGE
             characters, treat it as a scanned/image page and run OCR
             on that page only via PDFOCRService.
          3. If OCR is unavailable (no API key) or fails, skip the page
             with a warning rather than crashing the whole document.

        This means a mixed PDF (some text pages, some scanned pages)
        is handled correctly — each page is judged independently.
        """
        pages = []
        ocr_service: Optional[PDFOCRService] = None

        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        except Exception as e:
            raise RuntimeError(f"pymupdf failed to open '{source}': {e}") from e

        try:
            for page_num in range(len(doc)):
                page = doc[page_num]
                text = page.get_text("text").strip()

                if len(text) >= OCR_FALLBACK_MIN_CHARS_PER_PAGE:
                    # ✅ Good text extraction — use pymupdf result
                    pages.append({"page_number": page_num + 1, "text": text})
                    continue

                # ⚠️ Page has little/no text — likely scanned image
                logger.info(
                    "Page %d of '%s' has only %d chars from pymupdf — attempting OCR fallback",
                    page_num + 1, source, len(text),
                )

                # Lazy-init OCR service (only if actually needed)
                if ocr_service is None:
                    try:
                        ocr_service = PDFOCRService()
                    except Exception as ocr_init_err:
                        # OCR not configured (no API key) — skip this page
                        logger.warning(
                            "OCR service unavailable for page %d of '%s': %s — skipping page",
                            page_num + 1, source, ocr_init_err,
                        )
                        if text:
                            # Keep the little pymupdf text we do have
                            pages.append({"page_number": page_num + 1, "text": text})
                        continue

                # Render the single page to a PDF in memory and send to OCR
                try:
                    single_page_doc = fitz.open()
                    single_page_doc.insert_pdf(doc, from_page=page_num, to_page=page_num)
                    single_page_bytes = single_page_doc.tobytes()
                    single_page_doc.close()

                    ocr_text = ocr_service.extract_text_from_bytes(
                        single_page_bytes,
                        filename=f"{source}_page{page_num + 1}.pdf",
                    ).strip()

                    if ocr_text:
                        logger.info(
                            "✅ OCR extracted %d chars from page %d of '%s'",
                            len(ocr_text), page_num + 1, source,
                        )
                        pages.append({"page_number": page_num + 1, "text": ocr_text})
                    else:
                        logger.warning(
                            "OCR returned empty text for page %d of '%s' — skipping",
                            page_num + 1, source,
                        )

                except OCRError as ocr_err:
                    logger.warning(
                        "OCR failed for page %d of '%s': %s — skipping page",
                        page_num + 1, source, ocr_err,
                    )
                    if text:
                        pages.append({"page_number": page_num + 1, "text": text})

        finally:
            doc.close()
            if ocr_service is not None:
                ocr_service.close()

        if not pages:
            raise ValueError(
                f"No extractable text found in '{source}'. "
                "All pages failed both pymupdf extraction and OCR fallback."
            )
        return pages

    @staticmethod
    def _extract_csv(csv_bytes: bytes, source: str) -> List[dict]:
        """
        Parse CSV into a single text block.
        Each row → "col1: val1 | col2: val2 ..." line.
        page_number is None (no page concept for CSV).
        """
        try:
            text_content = csv_bytes.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text_content))
            rows = list(reader)
            if not rows:
                raise ValueError(f"CSV '{source}' is empty")

            lines = []
            for row in rows:
                line = " | ".join(
                    f"{k.strip()}: {v.strip()}" for k, v in row.items() if v and v.strip()
                )
                if line:
                    lines.append(line)

            if not lines:
                raise ValueError(f"No data rows found in CSV '{source}'")

            return [{"page_number": None, "text": "\n".join(lines)}]

        except Exception as e:
            raise RuntimeError(f"CSV parsing failed for '{source}': {e}") from e

    @staticmethod
    def _extract_xlsx(xlsx_bytes: bytes, source: str) -> List[dict]:
        """
        Parse XLSX — one section per sheet.
        Each row → space-separated cell values.
        page_number = sheet index (1-based).
        """
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(xlsx_bytes), read_only=True, data_only=True
            )
            pages = []
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]
                lines = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cells:
                        lines.append(" | ".join(cells))
                if lines:
                    pages.append({
                        "page_number": sheet_idx,
                        "text": f"[Sheet: {sheet_name}]\n" + "\n".join(lines),
                    })
            wb.close()

            if not pages:
                raise ValueError(f"No data found in XLSX '{source}'")
            return pages

        except Exception as e:
            raise RuntimeError(f"XLSX parsing failed for '{source}': {e}") from e

    # ------------------------------------------------------------------
    # Chunker — shared by all formats
    # ------------------------------------------------------------------

    @classmethod
    def _chunk_pages(cls, pages: List[dict], source: str) -> List[TextChunk]:
        """
        Chunk each page/section and preserve page_number mapping.
        Uses langchain RecursiveCharacterTextSplitter.
        """
        chunks: List[TextChunk] = []
        chunk_index = 0

        for page in pages:
            page_chunks = cls._splitter.split_text(page["text"])
            for raw_chunk in page_chunks:
                cleaned = raw_chunk.strip()
                if not cleaned:
                    continue
                chunks.append(TextChunk(
                    text=cleaned,
                    chunk_index=chunk_index,
                    page_number=page["page_number"],
                    source=source,
                ))
                chunk_index += 1

        return chunks